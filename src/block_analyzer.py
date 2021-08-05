#!/usr/bin/python3
import common
import psycopg2
from dateutil.parser import parse
import json
import datetime
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
def block_analyzer_peer_week():
    conn = None
    try:
        # read connection parameters
        params = common.config()
        # connect to the PostgreSQL server
        print('Connecting to the PostgreSQL database...')
        conn = psycopg2.connect(**params)

        # create a cursor
        cur = conn.cursor()
        sql='SELECT time,number,hash FROM aggron_block WHERE time > current_timestamp - interval \'7 day\' ORDER BY time ASC'
        cur.execute(sql)

        # Wirte the 7 days blocks info to a tmp file
        rows = cur.fetchall()
        data={}
        data['blocks']=[]
        for row in rows:
            strtime=datetime.datetime.strftime(row[0],"%Y-%m-%dT%H:%M:%S.%f")
            #   print(row[0])
            #   print(strtime)
            #   print(datetime.datetime.strptime(strtime,"%Y-%m-%dT%H:%M:%S.%f"))
            data['blocks'].append({
            'block_time':strtime,
            'block_number':row[1],
            'block_hash':row[2]
            })
        with open('block_data.txt', 'w') as outfile:
                json.dump(data, outfile) 
	# close the communication with the PostgreSQL
        cur.close()
    except (Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()
            print('Database connection closed.')

def generate_csv_file():
    block_analyzer_peer_week()
    f = open("block_data.txt","r") 
    data = json.load(f)
    wb = Workbook()
    ws = wb.active
    ws1= wb.create_sheet("Datas_info",0)
    # Sheet for data info detals
    ws.title = "block_intervals"
    ws['A1'] = "Block_I_hash"
    ws['B1'] = "Block_I+1_hash"
    ws['C1'] = "Block_I+2_hash"
    ws['D1'] = "Block_I_to(I+1)_interval(secondes)"
    ws['E1'] = "Block_I_to(I+2)_interval(secondes)"
    interval_count=0
    interval_2_count=0
    for num in range(len(data['blocks'])):
        block_A_number=data['blocks'][num]['block_number']
        block_A_hash=data['blocks'][num]['block_hash']
        block_A_time=data['blocks'][num]['block_time']
        block_B_number=""
        block_B_hash=""
        block_B_time=""
        block_C_number=""
        block_C_hash=""
        block_C_time=""
        block_interval=""
        block_2_interval=""
        i=num+1
        j=num+2
        if (len(data['blocks']) - i) > 0:
          block_B_number=data['blocks'][i]['block_number']
          block_B_hash=data['blocks'][i]['block_hash']
          block_B_time=data['blocks'][i]['block_time']
        if (len(data['blocks']) - j) > 0:
           block_C_number=data['blocks'][i]['block_number']
           block_C_hash=data['blocks'][i]['block_hash']
           block_C_time=data['blocks'][j]['block_time']
        if (len(block_A_time) != 0) & (len(block_B_time) != 0):
            block_interval=(datetime.datetime.strptime(block_B_time,"%Y-%m-%dT%H:%M:%S.%f") - datetime.datetime.strptime(block_A_time,"%Y-%m-%dT%H:%M:%S.%f")).total_seconds()
            if block_interval > 15:
                interval_count +=1
        if (len(block_A_time) != 0) & (len(block_C_time) != 0):
            block_2_interval=(datetime.datetime.strptime(block_C_time,"%Y-%m-%dT%H:%M:%S.%f") - datetime.datetime.strptime(block_A_time,"%Y-%m-%dT%H:%M:%S.%f")).total_seconds()
            if block_2_interval > 15:
                interval_2_count +=1 
        ws.append([block_A_hash, block_B_hash, block_C_hash,block_interval,block_2_interval])
        
    #data style formatting depends on percentile
    ws.conditional_formatting.add('D2:D'+str(ws.max_row),
                   ColorScaleRule(start_type='percentile', start_value=0, start_color='E3F6CE',
                   end_type='percentile', end_value=99.9, end_color='AA0000')
                            )
    ws.conditional_formatting.add('E2:E'+str(ws.max_row),
                   ColorScaleRule(start_type='percentile', start_value=0, start_color='E3F6CE',
                   end_type='percentile', end_value=99.9, end_color='AA0000')
                            )

     #Sheet for data analyzer
    interval_avg='=AVERAGE(block_intervals!D2:D'+str(ws.max_row)+')'
    interval_2_avg='=AVERAGE(block_intervals!E2:E'+str(ws.max_row)+')'
    Block_Interval_PERCENTILE_999='=PERCENTILE(block_intervals!D2:D'+str(ws.max_row)+',0.999)'
    Block_2_Interval_PERCENTILE_999='=PERCENTILE(block_intervals!E2:E'+str(ws.max_row)+',0.999)'
    Block_Interval_PERCENTILE_90='=PERCENTILE(block_intervals!D2:D'+str(ws.max_row)+',0.9)'
    Block_2_Interval_PERCENTILE_90='=PERCENTILE(block_intervals!E2:E'+str(ws.max_row)+',0.9)'
    Block_Interval_PERCENTILE_95='=PERCENTILE(block_intervals!D2:D'+str(ws.max_row)+',0.95)'
    Block_2_Interval_PERCENTILE_95='=PERCENTILE(block_intervals!E2:E'+str(ws.max_row)+',0.95)'
    Block_Interval_PERCENTILE_99='=PERCENTILE(block_intervals!D2:D'+str(ws.max_row)+',0.99)'
    Block_2_Interval_PERCENTILE_99='=PERCENTILE(block_intervals!E2:E'+str(ws.max_row)+',0.99)'
    Max_block_internal='=MAX(block_intervals!D2:D'+str(ws.max_row)+')'
    Max_2_block_internal='=MAX(block_intervals!E2:E'+str(ws.max_row)+')'

    ws1['A1'] = "Block_I_to(I+1)_interval_Avg"
    ws1['B1'] = interval_avg
    ws1['A2'] = "Block_I_to(I+1)_interval_Max"
    ws1['B2'] = Max_block_internal
    ws1['A3'] = "Block_Interval>15s"
    ws1['B3'] = interval_count
    ws1['A4'] = "Block_Interval_PERCENTILE_90"
    ws1['B4'] = Block_Interval_PERCENTILE_90
    ws1['A5'] = "Block_Interval_PERCENTILE_95"
    ws1['B5'] = Block_Interval_PERCENTILE_95
    ws1['A6'] = "Block_Interval_PERCENTILE_99"
    ws1['B6'] = Block_Interval_PERCENTILE_99
    ws1['A7'] = "Block_Interval_PERCENTILE_999"
    ws1['B7'] = Block_Interval_PERCENTILE_999

    ws1['A9'] = "2_Blocks_interval_Avg"
    ws1['B9'] = interval_2_avg
    ws1['A10'] = "2_Blocks_interval_Max"
    ws1['B10'] = Max_2_block_internal
    ws1['A11'] = "2_Blocks_Interval>30s"
    ws1['B11'] = interval_2_count
    ws1['A12'] = "2_Blocks_Interval_PERCENTILE_90"
    ws1['B12'] = Block_2_Interval_PERCENTILE_90
    ws1['A13'] = "2_Blocks_Interval_PERCENTILE_95"
    ws1['B13'] = Block_2_Interval_PERCENTILE_95
    ws1['A14'] = "2_Blocks_Interval_PERCENTILE_99"
    ws1['B14'] = Block_2_Interval_PERCENTILE_99
    ws1['A15'] = "2_Blocks_Interval_PERCENTILE_999"
    ws1['B15'] = Block_2_Interval_PERCENTILE_999

    wb.save("block_analyzer_"+datetime.datetime.now().strftime('%Y%m%d%H%M%S')+".xlsx")
if __name__ == '__main__':
    generate_csv_file()