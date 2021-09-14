#!/usr/bin/python3
import os
import requests
import json
import time
import datetime
import statsd
from openpyxl import Workbook
print(str(os.getenv('GITHUB_TOKEN')))
headers = {"Authorization": "token "+str(os.getenv('GITHUB_TOKEN'))}
data = {}
data['workflow_info']=[]
def run_query(url): # A simple function to use requests.post to make the API call. Note the json= section.
    request = requests.get(url, headers=headers)
    print(request)
    link = request.headers.get('link', None)
    if link is not None:
        print(link)
    if request.status_code == 200:
        return request.json()
    else:
        raise Exception("Query failed to run by returning code of {}. {}".format(request.status_code))
def generate_workflow_info(workflow_result):
    is_break=False
    for num in range(len(workflow_result["workflow_runs"])):
        if (workflow_result["workflow_runs"][num]["conclusion"] != "skipped") &  (workflow_result["workflow_runs"][num]["status"] != "queued" ) & (workflow_result["workflow_runs"][num]["status"] != "" ):
            start_time=datetime.datetime.strptime(workflow_result["workflow_runs"][num]["created_at"],"%Y-%m-%dT%H:%M:%SZ")
            a_week_ago=datetime.datetime.strptime((datetime.datetime.now()+datetime.timedelta(days=-7)).strftime('%Y-%m-%dT%H:%M:%SZ'),"%Y-%m-%dT%H:%M:%SZ")
            if a_week_ago <= start_time:
                # jobs=get_job_info(workflow_result["workflow_runs"][num]["jobs_url"])
                data['workflow_info'].append({
                    'workflow_id': workflow_result["workflow_runs"][num]["id"],
                    'workflow_status':workflow_result["workflow_runs"][num]["status"],
                    'workflow_conclusion':workflow_result["workflow_runs"][num]["conclusion"],
                    'workflow_html_url':workflow_result["workflow_runs"][num]["html_url"],
                    'workflow_created_at':workflow_result["workflow_runs"][num]["created_at"],
                    'workflow_updated_at':workflow_result["workflow_runs"][num]["updated_at"]
                    })
            else:
                is_break=True
                break
    return is_break
def generate_useful_info(workfllow_name):
    url = "https://api.github.com/repos/nervosnetwork/ckb/actions/workflows/"+workfllow_name+".yaml/runs?page=1&per_page=100"
    res=requests.get(url,headers=headers)
    generate_workflow_info(res.json())
    while 'next' in res.links.keys():
        res=requests.get(res.links['next']['url'],headers=headers)
        is_break=generate_workflow_info(res.json())
        if is_break is True:
            break
    with open('CI_result_data.txt', 'w') as outfile:
            json.dump(data, outfile)
def send_metrics(metrics_value,metrics_name):
    print(metrics_value)
    c = statsd.StatsClient(localhst, 8125)
    c.timing("github_actions.ckb."+metrics_name+".duration", metrics_value)
    # c.incr("github_actions.ckb."+metrics_name+".count",metrics_value)
    # c.gauge("github_actions.ckb."+metrics_name+".gauge", metrics_value, rate=1, delta=False)
def cellect_ci_ubuntu_metrics(workfllow_name):
    generate_useful_info(workfllow_name)
    f = open("CI_result_data.txt","r") 
    data = json.load(f)
    wb = Workbook()
    ws = wb.active
    ws.title = "workflow_runs_info"
    ws['A1'] = "Workflow_ID"
    ws['B1'] = "status"
    ws['C1'] = "workflow_duration"

    for i in range(len(data["workflow_info"])):
        start_time=datetime.datetime.strptime(data["workflow_info"][i]["workflow_created_at"],"%Y-%m-%dT%H:%M:%SZ")
        update_time=datetime.datetime.strptime(data["workflow_info"][i]["workflow_updated_at"],"%Y-%m-%dT%H:%M:%SZ")
        print("workflow name is"+workfllow_name)
        print("start time is "+str(start_time))
        print("update time is "+str(update_time))
        print("statsd start!"+str((update_time - start_time).total_seconds()*1000))
        ws.append([data["workflow_info"][i]["workflow_id"],data["workflow_info"][i]["workflow_status"],(update_time - start_time).total_seconds()*1000])
        send_metrics((update_time - start_time).total_seconds()*1000,workfllow_name)
    wb.save("weekly_CI_result_"+datetime.date.today().strftime('%Y%m%d')+".xlsx")
if __name__ == '__main__':
#   cellect_ci_ubuntu_metrics("ci_integration_tests_ubuntu")
#   time.sleep(120)
#   cellect_ci_ubuntu_metrics("ci_benchmarks_ubuntu")
#   time.sleep(120)
#   cellect_ci_ubuntu_metrics("ci_quick_checks_ubuntu")
#   time.sleep(120)
#   cellect_ci_ubuntu_metrics("ci_cargo_deny_ubuntu")
#   time.sleep(120)
#   cellect_ci_ubuntu_metrics("ci_linters_ubuntu")
#   time.sleep(120)
#   cellect_ci_ubuntu_metrics("ci_unit_tests_ubuntu")
#   time.sleep(120)
  cellect_ci_ubuntu_metrics("ci_wasm_build_ubuntu")


