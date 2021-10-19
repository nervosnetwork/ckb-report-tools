#!/usr/bin/python3
import requests
import json
import csv
import time
import datetime
import os
from openpyxl import Workbook
from dotenv import load_dotenv
load_dotenv()
headers = {"Authorization": "token "+str(os.getenv('GITHUB_TOKEN'))}
data = {}
data['workflow_info']=[]
def run_query(url): # A simple function to use requests.post to make the API call. Note the json= section.
    request = requests.get(url, headers=headers)
    link = request.headers.get('link', None)

    if link is not None:
        print(link)
    if request.status_code == 200:
        return request.json()
    else:
        raise Exception("Query failed to run by returning code of {}. {}".format(request.status_code))

def get_job_info(url):
    job_result = run_query(url)
    jobs = {}
    jobs['job_info']=[]
    for num in range(len(job_result["jobs"])):
        jobs["job_info"].append({
        'job_name':job_result["jobs"][num]['name'],
        'job_html_url':job_result["jobs"][num]['html_url'],
        'job_status':job_result["jobs"][num]['status'],
        'job_conclusion':job_result["jobs"][num]['conclusion'],
        'job_started_at':job_result["jobs"][num]['started_at'],
        'job_completed_at':job_result["jobs"][num]['completed_at']
        })
    return jobs

def generate_workflow_info(workflow_result):
    # workflow_result = run_query("https://api.github.com/repos/nervosnetwork/ckb/actions/workflows/ci.yaml/runs?per_page=100&status!='skipped'") # Execute the query
    is_break=False
    for num in range(len(workflow_result["workflow_runs"])):
        # print(workflow_result["workflow_runs"][num]["conclusion"])
        if (workflow_result["workflow_runs"][num]["conclusion"] != "skipped") &  (workflow_result["workflow_runs"][num]["status"] != "queued"):
            start_time=datetime.datetime.strptime(workflow_result["workflow_runs"][num]["created_at"],"%Y-%m-%dT%H:%M:%SZ")
            a_week_ago=datetime.datetime.strptime((datetime.datetime.now()+datetime.timedelta(days=-7)).strftime('%Y-%m-%dT%H:%M:%SZ'),"%Y-%m-%dT%H:%M:%SZ")
            print("Start time is"+str(start_time))
            print("Now time is"+str(a_week_ago))
            if a_week_ago <= start_time:
                jobs=get_job_info(workflow_result["workflow_runs"][num]["jobs_url"])
                data['workflow_info'].append({
                    'workflow_id': workflow_result["workflow_runs"][num]["id"],
                    'workflow_event': workflow_result["workflow_runs"][num]["event"],
                    'workflow_status':workflow_result["workflow_runs"][num]["status"],
                    'workflow_branch':workflow_result["workflow_runs"][num]["head_branch"],
                    'workflow_conclusion':workflow_result["workflow_runs"][num]["conclusion"],
                    'workflow_html_url':workflow_result["workflow_runs"][num]["html_url"],
                    'workflow_created_at':workflow_result["workflow_runs"][num]["created_at"],
                    'workflow_updated_at':workflow_result["workflow_runs"][num]["updated_at"],
                    'workflow_trigger_author':workflow_result["workflow_runs"][num]["head_commit"]["author"]["name"],
                    'jobs':jobs
                    })
            else:
                is_break=True
                break
    return is_break

def generate_useful_info():
    url = "https://api.github.com/repos/nervosnetwork/ckb/actions/workflows/ci.yaml/runs?page=1&per_page=100"
    res=requests.get(url,headers=headers)
    generate_workflow_info(res.json())
    while 'next' in res.links.keys():
        res=requests.get(res.links['next']['url'],headers={"Authorization": "token "+str(os.getenv('GITHUB_TOKEN'))})
        is_break=generate_workflow_info(res.json())
        if is_break is True:
            break
    with open('CI_result_data.txt', 'w') as outfile:
            json.dump(data, outfile)
def append_jobs_info_to_csv(workflow_id,jobs):
        jobs=jobs
        Linters_macOS_durations=""
        Linters_macOS_conclusion=""
        Linters_Linux_durations=""
        Linters_Linux_conclusion=""

        Quick_Check_durations=""
        Quick_Check_conclusion=""

        WASM_build_durations=""
        WASM_build_conclusion=""

        Security_Audit_Licenses_durations=""
        Security_Audit_Licenses_conclusion=""

        UnitTest_macOS_durations=""
        UnitTest_macOS_conclusion=""
        UnitTest_Linux_durations=""
        UnitTest_Linux_conclusion=""
        UnitTest_Windows_durations=""
        UnitTest_Windows_conclusion=""

        Integration_Test_macOS_durations=""
        Integration_Test_macOS_conclusion=""
        Integration_Test_Linux_durations=""
        Integration_Test_Linux_conclusion=""
        Integration_Test_Windows_durations=""
        Integration_Test_Windows_conclusion=""

        Benchmarks_Test_macOS_durations=""
        Benchmarks_Test_macOS_conclusion=""
        Benchmarks_Test_Linux_durations=""
        Benchmarks_Test_Linux_conclusion=""

        ci_durations=""
        ci_conclusion=""

        Check_Current_Workflow_durations=""
        Check_Current_Workflow_conclusione=""
        Check_Current_start_time=""
        Integration_test_max=""
        UnitTest_test_max=""
        Benchmarks_test_max=""
        for num in range(len(jobs["job_info"])):
            if (jobs["job_info"][num]["job_name"]).find("Linters (Linux)") != -1:
                print("Linters (Linux)")
                Linters_Linux_conclusion=jobs["job_info"][num]["job_conclusion"]
                Linters_Linux_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("Linters (macOS)") != -1:
                print("Linters (macOS)")
                Linters_macOS_conclusion=jobs["job_info"][num]["job_conclusion"]
                Linters_macOS_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("Quick_Check" ) != -1:
                print("Quick_Check")
                Quick_Check_conclusion=jobs["job_info"][num]["job_conclusion"]
                Quick_Check_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("WASM_build" ) != -1:
                print("WASM_build")
                WASM_build_conclusion=jobs["job_info"][num]["job_conclusion"]
                WASM_build_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("Security_Audit_Licenses" ) != -1:
                print("Security_Audit_Licenses")
                Security_Audit_Licenses_conclusion=jobs["job_info"][num]["job_conclusion"]
                Security_Audit_Licenses_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("Integration_Test (macOS)" ) != -1:
                print("Integration_Test (macOS)")
                Integration_Test_macOS_conclusion=jobs["job_info"][num]["job_conclusion"]
                Integration_Test_macOS_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("Integration_Test (Linux)" ) != -1:
                print("Integration_Test (Linux)")
                Integration_Test_Linux_conclusion=jobs["job_info"][num]["job_conclusion"]
                Integration_Test_Linux_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("Integration_Test (Windows)" ) != -1:
                print("Integration_Test (Windows)")
                Integration_Test_Windows_conclusion=jobs["job_info"][num]["job_conclusion"]
                Integration_Test_Windows_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("UnitTest (macOS)" ) != -1:
                print("UnitTest (macOS)")
                UnitTest_macOS_conclusion=jobs["job_info"][num]["job_conclusion"]
                UnitTest_macOS_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("UnitTest (Linux)" ) != -1:
                print("UnitTest (Linux)")
                UnitTest_Linux_conclusion=jobs["job_info"][num]["job_conclusion"]
                UnitTest_Linux_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("UnitTest (Windows)" ) != -1:
                print("UnitTest (Windows)")
                UnitTest_Windows_conclusion=jobs["job_info"][num]["job_conclusion"]
                UnitTest_Windows_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("Benchmarks_Test (macOS)" ) != -1:
                print("Benchmarks_Test (macOS)")
                Benchmarks_Test_macOS_conclusion=jobs["job_info"][num]["job_conclusion"]
                Benchmarks_Test_macOS_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("Benchmarks_Test (Linux)" ) != -1:
                print("Benchmarks_Test (Linux)")
                Benchmarks_Test_Linux_conclusion=jobs["job_info"][num]["job_conclusion"]
                Benchmarks_Test_Linux_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("ci" ) != -1:
                print("ci")
                ci_conclusion=jobs["job_info"][num]["job_conclusion"]
                ci_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
            if (jobs["job_info"][num]["job_name"]).find("Check If Current" ) != -1:
                print("Check If Current")
                Check_Current_Workflow_conclusione=jobs["job_info"][num]["job_conclusion"]
                Check_Current_start_time=jobs["job_info"][num]["job_started_at"]
                Check_Current_Workflow_durations=(datetime.datetime.strptime(jobs["job_info"][num]["job_completed_at"],"%Y-%m-%dT%H:%M:%SZ") - datetime.datetime.strptime(jobs["job_info"][num]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")).total_seconds()
        Integration_list=[Integration_Test_macOS_durations,Integration_Test_Linux_durations,Integration_Test_Windows_durations]
        Integration_test_max=max(Integration_list)
        UnitTest_list=[UnitTest_macOS_durations,UnitTest_Linux_durations,UnitTest_Windows_durations]
        UnitTest_test_max=max(UnitTest_list)
        Benchmarks_list=[Benchmarks_Test_macOS_durations,Benchmarks_Test_Linux_durations]
        Benchmarks_test_max=max(Benchmarks_list)
        Liners_list=[Linters_Linux_durations,Linters_macOS_durations]
        Liners_max=max(Liners_list)
        jobs_data=[workflow_id,Check_Current_Workflow_durations,Check_Current_Workflow_conclusione,Linters_macOS_durations,Linters_macOS_conclusion,Linters_Linux_durations,Linters_Linux_conclusion,Quick_Check_durations,Quick_Check_conclusion,WASM_build_durations,WASM_build_conclusion,Security_Audit_Licenses_durations,Security_Audit_Licenses_conclusion,UnitTest_macOS_durations,UnitTest_macOS_conclusion,UnitTest_Linux_durations,UnitTest_Linux_conclusion,UnitTest_Windows_durations,UnitTest_Windows_conclusion,Integration_Test_macOS_durations,Integration_Test_macOS_conclusion,Integration_Test_Linux_durations,Integration_Test_Linux_conclusion,Integration_Test_Windows_durations,Integration_Test_Windows_conclusion,Benchmarks_Test_macOS_durations,Benchmarks_Test_macOS_conclusion,Benchmarks_Test_Linux_durations,Benchmarks_Test_Linux_conclusion,ci_durations,ci_conclusion,Check_Current_start_time,Integration_test_max,UnitTest_test_max,Benchmarks_test_max,Liners_max]
        return jobs_data
def generate_csv_file():
    generate_useful_info()
    f = open("CI_result_data.txt","r") 
    data = json.load(f)
    wb = Workbook()
    ws = wb.active
    ws.title = "workflow_runs_info"
    ws['A1'] = "Workflow_ID"
    ws['B1'] = "workflow_start"
    ws['C1'] = "workflow_event"
    ws['D1'] = "author"
    ws['E1'] = "branch"
    ws['F1'] = "workflow_conclusion"
    ws['G1'] = "workflow_durations"
    ws['H1'] = "workflow_pending_time"
    ws['I1'] = "Jobs_execion_time_avg"
    ws['J1'] = "Jobs_execion_time_Max"
    ws['K1'] = "Jobs_execion_time_SUM"
    #Sheet for jobs_info
    jobs_sheet=wb.create_sheet("jobs_info")
    jobs_sheet['A1'] = "Workflow_ID"
    jobs_sheet['B1'] = "Check_Current_Workflow_durations"
    jobs_sheet['C1'] = "Check_Current_Workflow_conclusione"
    jobs_sheet['D1'] = "Linters_macOS_durations"
    jobs_sheet['E1'] = "Linters_macOS_conclusion"
    jobs_sheet['F1'] = "Linters_Linux_durations"
    jobs_sheet['G1'] = "Linters_Linux_conclusion"
    jobs_sheet['H1'] = "Quick_Check_durations"
    jobs_sheet['I1'] = "Quick_Check_conclusion"
    jobs_sheet['J1'] = "WASM_build_durations"
    jobs_sheet['K1'] = "WASM_build_conclusion"
    jobs_sheet['L1'] = "Security_Audit_Licenses_durations"
    jobs_sheet['M1'] = "Security_Audit_Licenses_conclusion"
    jobs_sheet['N1'] = "UnitTest_macOS_durations"
    jobs_sheet['O1'] = "UnitTest_macOS_conclusion"
    jobs_sheet['P1'] = "UnitTest_Linux_durations"
    jobs_sheet['Q1'] = "UnitTest_Linux_conclusion"
    jobs_sheet['R1'] = "UnitTest_Windows_durations"
    jobs_sheet['S1'] = "UnitTest_Windows_conclusion"
    jobs_sheet['T1'] = "Integration_Test_macOS_durations"
    jobs_sheet['U1'] = "Integration_Test_macOS_conclusion"
    jobs_sheet['V1'] = "Integration_Test_Linux_durations"
    jobs_sheet['W1'] = "Integration_Test_Linux_conclusion"
    jobs_sheet['X1'] = "Integration_Test_Windows_durations"
    jobs_sheet['Y1'] = "Integration_Test_Windows_conclusion"
    jobs_sheet['Z1'] = "Benchmarks_Test_macOS_durations"
    jobs_sheet['AA1'] = "Benchmarks_Test_macOS_conclusion"
    jobs_sheet['AB1'] = "Benchmarks_Test_Linux_durations"
    jobs_sheet['AC1'] = "Benchmarks_Test_Linux_conclusion"
    jobs_sheet['AD1'] = "ci_durations"
    jobs_sheet['AE1'] = "ci_conclusion"
    jobs_sheet['AF1'] = "Check_Current_start_time"
    jobs_sheet['AG1'] = "Integration_test_max"
    jobs_sheet['AH1'] = "UnitTest_test_max"
    jobs_sheet['AI1'] = "Benchmarks_test_max"
    jobs_sheet['AJ1'] = "Liners_max"
    for i in range(len(data["workflow_info"])):
        start_time=datetime.datetime.strptime(data["workflow_info"][i]["workflow_created_at"],"%Y-%m-%dT%H:%M:%SZ")
        update_time=datetime.datetime.strptime(data["workflow_info"][i]["workflow_updated_at"],"%Y-%m-%dT%H:%M:%SZ")
        author=data["workflow_info"][i]["workflow_trigger_author"]
        jobs=data["workflow_info"][i]["jobs"]
        jobs_data=append_jobs_info_to_csv(data["workflow_info"][i]["workflow_id"],jobs)
        # apeend job info to jobs_data sheet
        jobs_sheet.append(jobs_data)
        first_job_start_time=""
        workflow_pending_time=""
        Jobs_execion_time_avg='=AVERAGE(jobs_info!AD'+str(ws.max_row+1)+',jobs_info!AB'+str(ws.max_row+1)+',jobs_info!Z'+str(ws.max_row+1)+',jobs_info!X'+str(ws.max_row+1)+',jobs_info!V'+str(ws.max_row+1)+',jobs_info!T'+str(ws.max_row+1)+',jobs_info!R'+str(ws.max_row+1)+',jobs_info!P'+str(ws.max_row+1)+',jobs_info!N'+str(ws.max_row+1)+',jobs_info!L'+str(ws.max_row+1)+',jobs_info!J'+str(ws.max_row+1)+',jobs_info!H'+str(ws.max_row+1)+',jobs_info!F'+str(ws.max_row+1)+',jobs_info!D'+str(ws.max_row+1)+',jobs_info!B'+str(ws.max_row+1)+')'
        Jobs_execion_time_Max='=MAX(jobs_info!AD'+str(ws.max_row+1)+',jobs_info!AB'+str(ws.max_row+1)+',jobs_info!Z'+str(ws.max_row+1)+',jobs_info!X'+str(ws.max_row+1)+',jobs_info!V'+str(ws.max_row+1)+',jobs_info!T'+str(ws.max_row+1)+',jobs_info!R'+str(ws.max_row+1)+',jobs_info!P'+str(ws.max_row+1)+',jobs_info!N'+str(ws.max_row+1)+',jobs_info!L'+str(ws.max_row+1)+',jobs_info!J'+str(ws.max_row+1)+',jobs_info!H'+str(ws.max_row+1)+',jobs_info!F'+str(ws.max_row+1)+',jobs_info!D'+str(ws.max_row+1)+',jobs_info!B'+str(ws.max_row+1)+')'
        Jobs_execion_time_SUM='=SUM(jobs_info!AD'+str(ws.max_row+1)+',jobs_info!AG'+str(ws.max_row+1)+',jobs_info!AH'+str(ws.max_row+1)+',jobs_info!AI'+str(ws.max_row+1)+',jobs_info!AJ'+str(ws.max_row+1)+',jobs_info!L'+str(ws.max_row+1)+',jobs_info!J'+str(ws.max_row+1)+',jobs_info!H'+str(ws.max_row+1)+',jobs_info!B'+str(ws.max_row+1)+')'
        if len(jobs['job_info']) != 0:
            first_job_start_time=datetime.datetime.strptime(jobs["job_info"][1]["job_started_at"],"%Y-%m-%dT%H:%M:%SZ")
            print("first_job_start_time: "+str(first_job_start_time))
            workflow_pending_time=(first_job_start_time - start_time).total_seconds()
            print("================")
            print("workflow_pending_time: "+str(workflow_pending_time))
        ws.append([data["workflow_info"][i]["workflow_id"],start_time,data["workflow_info"][i]["workflow_event"],author,data["workflow_info"][i]["workflow_branch"],data["workflow_info"][i]["workflow_conclusion"],(update_time - start_time).total_seconds(),workflow_pending_time,Jobs_execion_time_avg,Jobs_execion_time_Max,Jobs_execion_time_SUM])
    
    # sheet for data analyze
    workflow_run_avg='=AVERAGE(workflow_runs_info!G2:G'+str(ws.max_row)+')/'+str(60)
    # workflow_percentile_99='=PERCENTILE(workflow_runs_info!G2:G'+str(ws.max_row)+',0.99)'
    workflow_success_percentile_99='=PERCENTILE(IF(workflow_runs_info!F2:F'+str(ws.max_row)+'="success",workflow_runs_info!G2:G'+str(ws.max_row)+'),0.99)/'+str(60)
    workflow_max='=MAX(workflow_runs_info!G2:G'+str(ws.max_row)+')/'+str(60)
    workflow_success_rate='=COUNTIF(workflow_runs_info!F2:F'+str(ws.max_row)+',"success")/'+str(ws.max_row-1)+'*100'
    

    workflow_wait_avg='=AVERAGE(workflow_runs_info!H2:H'+str(ws.max_row)+')/'+str(60)
    # workflow_wait_percentile_99='=PERCENTILE(workflow_runs_info!H2:H'+str(ws.max_row)+',0.99)'
    workflow_success_wait_percentile_99='=PERCENTILE(IF(workflow_runs_info!F2:F'+str(ws.max_row)+'="success",workflow_runs_info!H2:H'+str(ws.max_row)+'),0.99)/'+str(60)
    workflow_wait_max='=MAX(workflow_runs_info!H2:H'+str(ws.max_row)+')/'+str(60)

    Check_Current_Workflow_durations_avg='=AVERAGE(jobs_info!B2:B'+str(jobs_sheet.max_row)+')/'+str(60)
    # Check_Current_Workflow_durations_percentile_99='=PERCENTILE(jobs_info!B2:B'+str(jobs_sheet.max_row)+',0.99)'
    Check_Current_Workflow_durations_percentile_99='=PERCENTILE(IF(jobs_info!C2:C'+str(jobs_sheet.max_row)+'="success",jobs_info!B2:B'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    Check_Current_Workflow_durations_max='=MAX(jobs_info!B2:B'+str(jobs_sheet.max_row)+')/'+str(60)

    Linters_macOS_durations_avg='=AVERAGE(jobs_info!D2:D'+str(jobs_sheet.max_row)+')/'+str(60)
    Linters_macOS_percentile_99='=PERCENTILE(IF(jobs_info!E2:E'+str(jobs_sheet.max_row)+'="success",jobs_info!D2:D'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    Linters_macOS_max='=MAX(jobs_info!D2:D'+str(jobs_sheet.max_row)+')/'+str(60)
    Linters_macOS_success_rate='=COUNTIF(jobs_info!E2:E'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'

    Linters_Linux_durations_avg='=AVERAGE(jobs_info!F2:F'+str(jobs_sheet.max_row)+')/'+str(60)
    Linters_Linux_percentile_99='=PERCENTILE(IF(jobs_info!G2:G'+str(jobs_sheet.max_row)+'="success",jobs_info!F2:F'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    Linters_Linux_max='=MAX(jobs_info!F2:F'+str(jobs_sheet.max_row)+')/'+str(60)
    Linters_Linux_success_rate='=COUNTIF(jobs_info!G2:G'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'
	
    Quick_Check_durations_avg='=AVERAGE(jobs_info!H2:H'+str(jobs_sheet.max_row)+')/'+str(60)
    Quick_Check_percentile_99='=PERCENTILE(IF(jobs_info!I2:I'+str(jobs_sheet.max_row)+'="success",jobs_info!H2:H'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    Quick_Check_max='=MAX(jobs_info!H2:H'+str(jobs_sheet.max_row)+')/'+str(60)
    Quick_Check_success_rate='=COUNTIF(jobs_info!I2:I'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'
	
    WASM_build_durations_avg='=AVERAGE(jobs_info!J2:J'+str(jobs_sheet.max_row)+')/'+str(60)
    WASM_build_percentile_99='=PERCENTILE(IF(jobs_info!K2:K'+str(jobs_sheet.max_row)+'="success",jobs_info!J2:J'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    WASM_build_max='=MAX(jobs_info!J2:J'+str(jobs_sheet.max_row)+')/'+str(60)
    WASM_build_success_rate='=COUNTIF(jobs_info!K2:K'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'
	
    Security_Audit_Licenses_durations_avg='=AVERAGE(jobs_info!L2:L'+str(jobs_sheet.max_row)+')/'+str(60)
    Security_Audit_Licenses_percentile_99='=PERCENTILE(IF(jobs_info!M2:M'+str(jobs_sheet.max_row)+'="success",jobs_info!L2:L'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    Security_Audit_Licenses_max='=MAX(jobs_info!L2:L'+str(jobs_sheet.max_row)+')/'+str(60)
    Security_Audit_success_rate='=COUNTIF(jobs_info!M2:M'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'
	
    UnitTest_macOS_durations_avg='=AVERAGE(jobs_info!N2:N'+str(jobs_sheet.max_row)+')/'+str(60)
    UnitTest_macOS_percentile_99='=PERCENTILE(IF(jobs_info!O2:O'+str(jobs_sheet.max_row)+'="success",jobs_info!N2:N'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    UnitTest_macOS_max='=MAX(jobs_info!N2:N'+str(jobs_sheet.max_row)+')/'+str(60)
    UnitTest_macOS_success_rate='=COUNTIF(jobs_info!O2:O'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'
	
    UnitTest_Linux_durations_avg='=AVERAGE(jobs_info!P2:P'+str(jobs_sheet.max_row)+')/'+str(60)
    UnitTest_Linux_percentile_99='=PERCENTILE(IF(jobs_info!Q2:Q'+str(jobs_sheet.max_row)+'="success",jobs_info!P2:P'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    UnitTest_Linux_max='=MAX(jobs_info!P2:P'+str(jobs_sheet.max_row)+')/'+str(60)
    UnitTest_Linux_success_rate='=COUNTIF(jobs_info!Q2:Q'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'
	
    UnitTest_Windows_durations_avg='=AVERAGE(jobs_info!R2:R'+str(jobs_sheet.max_row)+')/'+str(60)
    UnitTest_Windows_percentile_99='=PERCENTILE(IF(jobs_info!S2:S'+str(jobs_sheet.max_row)+'="success",jobs_info!R2:R'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    UnitTest_Windows_max='=MAX(jobs_info!R2:R'+str(jobs_sheet.max_row)+')/'+str(60)
    UnitTest_Windows_success_rate='=COUNTIF(jobs_info!S2:S'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'
	
    Integration_Test_macOS_durations_avg='=AVERAGE(jobs_info!T2:T'+str(jobs_sheet.max_row)+')/'+str(60)
    Integration_Test_macOS_percentile_99='=PERCENTILE(IF(jobs_info!U2:U'+str(jobs_sheet.max_row)+'="success",jobs_info!T2:T'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    Integration_Test_macOS_max='=MAX(jobs_info!T2:T'+str(jobs_sheet.max_row)+')/'+str(60)
    Integration_Test_macOS_success_rate='=COUNTIF(jobs_info!U2:U'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'
	
    Integration_Test_Linux_durations_avg='=AVERAGE(jobs_info!V2:V'+str(jobs_sheet.max_row)+')/'+str(60)
    Integration_Test_Linux_percentile_99='=PERCENTILE(IF(jobs_info!W2:W'+str(jobs_sheet.max_row)+'="success",jobs_info!V2:V'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    Integration_Test_Linux_max='=MAX(jobs_info!V2:V'+str(jobs_sheet.max_row)+')/'+str(60)
    Integration_Test_Linux_success_rate='=COUNTIF(jobs_info!W2:W'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'
	
    Integration_Test_Windows_durations_avg='=AVERAGE(jobs_info!X2:X'+str(jobs_sheet.max_row)+')/'+str(60)
    Integration_Test_Windows_percentile_99='=PERCENTILE(IF(jobs_info!Y2:Y'+str(jobs_sheet.max_row)+'="success",jobs_info!X2:X'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    Integration_Test_Windows_max='=MAX(jobs_info!X2:X'+str(jobs_sheet.max_row)+')/'+str(60)
    Integration_Test_Windows_success_rate='=COUNTIF(jobs_info!Y2:Y'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'
	
    Benchmarks_Test_macOS_durations_avg='=AVERAGE(jobs_info!Z2:Z'+str(jobs_sheet.max_row)+')/'+str(60)
    Benchmarks_Test_macOS_percentile_99='=PERCENTILE(IF(jobs_info!AA2:AA'+str(jobs_sheet.max_row)+'="success",jobs_info!Z2:Z'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    Benchmarks_Test_macOS_max='=MAX(jobs_info!Z2:Z'+str(jobs_sheet.max_row)+')/'+str(60)
    Benchmarks_Test_macOS_success_rate='=COUNTIF(jobs_info!AA2:AA'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'
	
    Benchmarks_Test_Linux_durations_avg='=AVERAGE(jobs_info!AB2:AB'+str(jobs_sheet.max_row)+')/'+str(60)
    Benchmarks_Test_Linux_percentile_99='=PERCENTILE(IF(jobs_info!AC2:AC'+str(jobs_sheet.max_row)+'="success",jobs_info!AB2:AB'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    Benchmarks_Test_Linux_max='=MAX(jobs_info!AB2:AB'+str(jobs_sheet.max_row)+')/'+str(60)
    Benchmarks_Test_Linux_success_rate='=COUNTIF(jobs_info!AC2:AC'+str(jobs_sheet.max_row)+',"success")/'+str(jobs_sheet.max_row-1)+'*100'
	
    ci_durations_avg='=AVERAGE(jobs_info!AD2:AD'+str(jobs_sheet.max_row)+')/'+str(60)
    ci_durations_percentile_99='=PERCENTILE(IF(jobs_info!AE2:AE'+str(jobs_sheet.max_row)+'="success",jobs_info!AD2:AD'+str(jobs_sheet.max_row)+'),0.99)/'+str(60)
    ci_durations_max='=MAX(jobs_info!AD2:AD'+str(jobs_sheet.max_row)+')/'+str(60)

    data_analyze=wb.create_sheet("data_analyze",0)
    data_analyze['A1'] = "workflow_run_avg(minues)"
    data_analyze['B1'] = workflow_run_avg
    data_analyze['A2'] = "workflow_success_percentile_99(minues)"
    data_analyze['B2'] = workflow_success_percentile_99
    data_analyze['A3'] = "workflow_max(minues)"
    data_analyze['B3'] = workflow_max
    data_analyze['A4'] = "workflow_success_rate(%)"
    data_analyze['B4'] = workflow_success_rate

    data_analyze['A6'] = "workflow_wait_avg(minues)"
    data_analyze['B6'] = workflow_wait_avg
    data_analyze['A7'] = "workflow_success_wait_percentile_99(minues)"
    data_analyze['B7'] = workflow_success_wait_percentile_99
    data_analyze['A8'] = "workflow_wait_max(minues)"
    data_analyze['B8'] = workflow_wait_max


    data_analyze['A10'] = "Check_Current_Workflow_durations_avg(minues)"
    data_analyze['B10'] = Check_Current_Workflow_durations_avg
    data_analyze['A11'] = "Check_Current_Workflow_durations_percentile_99(minues)"
    data_analyze['B11'] = Check_Current_Workflow_durations_percentile_99
    data_analyze['A12'] = "Check_Current_Workflow_durations_max(minues)"
    data_analyze['B12'] = Check_Current_Workflow_durations_max

    data_analyze['A14'] = "Linters_macOS_durations_avg(minues)"
    data_analyze['B14'] = Linters_macOS_durations_avg
    data_analyze['A15'] = "Linters_macOS_percentile_99(minues)"
    data_analyze['B15'] = Linters_macOS_percentile_99
    data_analyze['A16'] = "Linters_macOS_max(minues)"
    data_analyze['B16'] = Linters_macOS_max
    data_analyze['A17'] = "Linters_macOS_success_rate"
    data_analyze['B17'] = Linters_macOS_success_rate

    data_analyze['A19'] = "Linters_Linux_durations_avg(minues)"
    data_analyze['B19'] = Linters_Linux_durations_avg
    data_analyze['A20'] = "Linters_Linux_percentile_99(minues)"
    data_analyze['B20'] = Linters_Linux_percentile_99
    data_analyze['A21'] = "Linters_Linux_max(minues)"
    data_analyze['B21'] = Linters_Linux_max
    data_analyze['A22'] = "Linters_Linux_success_rate"
    data_analyze['B22'] = Linters_Linux_success_rate

    data_analyze['A24'] = "Quick_Check_durations_avg(minues)"
    data_analyze['B24'] = Quick_Check_durations_avg
    data_analyze['A25'] = "Quick_Check_percentile_99(minues)"
    data_analyze['B25'] = Quick_Check_percentile_99
    data_analyze['A26'] = "Quick_Check_max(minues)"
    data_analyze['B26'] = Quick_Check_max
    data_analyze['A27'] = "Quick_Check_success_rate"
    data_analyze['B27'] = Quick_Check_success_rate

    data_analyze['A29'] = "WASM_build_durations_avg(minues)"
    data_analyze['B29'] = WASM_build_durations_avg
    data_analyze['A30'] = "WASM_build_percentile_99(minues)"
    data_analyze['B30'] = WASM_build_percentile_99
    data_analyze['A31'] = "WASM_build_max(minues)"
    data_analyze['B31'] = WASM_build_max
    data_analyze['A32'] = "WASM_build_success_rate"
    data_analyze['B32'] = WASM_build_success_rate

    data_analyze['A34'] = "Security_Audit_Licenses_durations_avg(minues)"
    data_analyze['B34'] = Security_Audit_Licenses_durations_avg
    data_analyze['A35'] = "Security_Audit_Licenses_percentile_99(minues)"
    data_analyze['B35'] = Security_Audit_Licenses_percentile_99
    data_analyze['A36'] = "Security_Audit_Licenses_max(minues)"
    data_analyze['B36'] = Security_Audit_Licenses_max
    data_analyze['A37'] = "Security_Audit_success_rate"
    data_analyze['B37'] = Security_Audit_success_rate

    data_analyze['A39'] = "UnitTest_macOS_durations_avg(minues)"
    data_analyze['B39'] = UnitTest_macOS_durations_avg
    data_analyze['A40'] = "UnitTest_macOS_percentile_99(minues)"
    data_analyze['B40'] = UnitTest_macOS_percentile_99
    data_analyze['A41'] = "UnitTest_macOS_max(minues)"
    data_analyze['B41'] = UnitTest_macOS_max
    data_analyze['A42'] = "UnitTest_macOS_success_rate"
    data_analyze['B42'] = UnitTest_macOS_success_rate

    data_analyze['A44'] = "UnitTest_Linux_durations_avg(minues)"
    data_analyze['B44'] = UnitTest_Linux_durations_avg
    data_analyze['A45'] = "UnitTest_Linux_percentile_99(minues)"
    data_analyze['B45'] = UnitTest_Linux_percentile_99
    data_analyze['A46'] = "UnitTest_Linux_max(minues)"
    data_analyze['B46'] = UnitTest_Linux_max
    data_analyze['A47'] = "UnitTest_Linux_success_rate"
    data_analyze['B47'] = UnitTest_Linux_success_rate

    data_analyze['A49'] = "UnitTest_Windows_durations_avg(minues)"
    data_analyze['B49'] = UnitTest_Windows_durations_avg
    data_analyze['A50'] = "UnitTest_Windows_percentile_99(minues)"
    data_analyze['B50'] = UnitTest_Windows_percentile_99
    data_analyze['A51'] = "UnitTest_Windows_max(minues)"
    data_analyze['B51'] = UnitTest_Windows_max
    data_analyze['A52'] = "UnitTest_Windows_success_rate"
    data_analyze['B52'] = UnitTest_Windows_success_rate

    data_analyze['A54'] = "Integration_Test_macOS_durations_avg(minues)"
    data_analyze['B54'] = Integration_Test_macOS_durations_avg
    data_analyze['A55'] = "Integration_Test_macOS_percentile_99(minues)"
    data_analyze['B55'] = Integration_Test_macOS_percentile_99
    data_analyze['A56'] = "Integration_Test_macOS_max(minues)"
    data_analyze['B56'] = Integration_Test_macOS_max
    data_analyze['A57'] = "Integration_Test_macOS_success_rate"
    data_analyze['B57'] = Integration_Test_macOS_success_rate

    data_analyze['A59'] = "Integration_Test_Linux_durations_avg(minues)"
    data_analyze['B59'] = Integration_Test_Linux_durations_avg
    data_analyze['A60'] = "Integration_Test_Linux_percentile_99(minues)"
    data_analyze['B60'] = Integration_Test_Linux_percentile_99
    data_analyze['A61'] = "Integration_Test_Linux_max(minues)"
    data_analyze['B61'] = Integration_Test_Linux_max
    data_analyze['A62'] = "Integration_Test_Linux_success_rate"
    data_analyze['B62'] = Integration_Test_Linux_success_rate

    data_analyze['A64'] = "Integration_Test_Windows_durations_avg(minues)"
    data_analyze['B64'] = Integration_Test_Windows_durations_avg
    data_analyze['A65'] = "Integration_Test_Windows_percentile_99(minues)"
    data_analyze['B65'] = Integration_Test_Windows_percentile_99
    data_analyze['A66'] = "Integration_Test_Windows_max(minues)"
    data_analyze['B66'] = Integration_Test_Windows_max
    data_analyze['A67'] = "Integration_Test_Windows_success_rate"
    data_analyze['B67'] = Integration_Test_Windows_success_rate

    data_analyze['A69'] = "Benchmarks_Test_macOS_durations_avg(minues)"
    data_analyze['B69'] = Benchmarks_Test_macOS_durations_avg
    data_analyze['A70'] = "Benchmarks_Test_macOS_percentile_99(minues)"
    data_analyze['B70'] = Benchmarks_Test_macOS_percentile_99
    data_analyze['A71'] = "Benchmarks_Test_macOS_max(minues)"
    data_analyze['B71'] = Benchmarks_Test_macOS_max
    data_analyze['A72'] = "Benchmarks_Test_macOS_success_rate"
    data_analyze['B72'] = Benchmarks_Test_macOS_success_rate

    data_analyze['A73'] = "Benchmarks_Test_Linux_durations_avg(minues)"
    data_analyze['B73'] = Benchmarks_Test_Linux_durations_avg
    data_analyze['A74'] = "Benchmarks_Test_Linux_percentile_99(minues)"
    data_analyze['B74'] = Benchmarks_Test_Linux_percentile_99
    data_analyze['A75'] = "Benchmarks_Test_Linux_max(minues)"
    data_analyze['B75'] = Benchmarks_Test_Linux_max
    data_analyze['A76'] = "Benchmarks_Test_Linux_success_rate"
    data_analyze['B76'] = Benchmarks_Test_Linux_success_rate

    data_analyze['A78'] = "ci_durations_avg(minues)"
    data_analyze['B78'] = ci_durations_avg
    data_analyze['A79'] = "ci_durations_percentile_99(minues)"
    data_analyze['B79'] = ci_durations_percentile_99
    data_analyze['A80'] = "ci_durations_max(minues)"
    data_analyze['B80'] = ci_durations_max

    wb.save("weekly_CI_result_"+datetime.date.today().strftime('%Y%m%d')+".xlsx")
if __name__ == '__main__':
#   generate_useful_info()
  generate_csv_file()

